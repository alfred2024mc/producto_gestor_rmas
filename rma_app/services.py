from dataclasses import replace
from datetime import datetime
from pathlib import Path
import re
import unicodedata

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd

from rma_app.config import ConfiguracionApp
from rma_app.models import EstadoBusqueda, ResultadoBusqueda


def normalizar_texto(valor: object) -> str:
    # Convertimos cualquier valor a texto.
    texto = str(valor)
    # Quitamos espacios al inicio y al final.
    texto_sin_bordes = texto.strip()
    # Convertimos a mayusculas para comparar mejor.
    texto_mayusculas = texto_sin_bordes.upper()
    # Quitamos espacios internos para tolerar diferencias de captura.
    texto_normalizado = texto_mayusculas.replace(" ", "")
    # Regresamos el valor listo para comparar.
    return texto_normalizado


def normalizar_nombre_columna(valor: object) -> str:
    # Homologamos encabezados con o sin acentos para tolerar variaciones del Excel.
    texto = unicodedata.normalize("NFKD", str(valor).strip().upper())
    texto = "".join(caracter for caracter in texto if not unicodedata.combining(caracter))
    return " ".join(texto.split())


def limpiar_valor_excel(valor: object) -> str:
    # Evitamos mostrar valores pandas vacios como texto "nan".
    if pd.isna(valor):
        return ""
    texto = str(valor).strip()
    return "" if texto.upper() == "NAN" else texto


def extraer_series(valor: object) -> list[str]:
    # Separamos una celda que puede traer varias series juntas.
    texto = str(valor).strip()
    if not texto:
        return []

    partes = re.split(r"[,;\n\r\t]+", texto)
    series = [normalizar_texto(parte) for parte in partes if normalizar_texto(parte)]

    # Si no hubo separadores validos, regresamos el valor normalizado como una sola serie.
    if not series:
        serie_unica = normalizar_texto(texto)
        return [serie_unica] if serie_unica else []

    return series


def parsear_fecha_texto(valor: object) -> datetime | None:
    # Intentamos convertir distintas representaciones comunes de fecha a datetime.
    texto = str(valor).strip()
    if not texto or texto.upper() == "NAN":
        return None

    fecha = pd.to_datetime(texto, errors="coerce")
    if pd.isna(fecha):
        return None

    return fecha.to_pydatetime()


def nombre_mes_espanol(fecha: datetime) -> str:
    meses = {
        1: "ENERO",
        2: "FEBRERO",
        3: "MARZO",
        4: "ABRIL",
        5: "MAYO",
        6: "JUNIO",
        7: "JULIO",
        8: "AGOSTO",
        9: "SEPTIEMBRE",
        10: "OCTUBRE",
        11: "NOVIEMBRE",
        12: "DICIEMBRE",
    }
    return meses[fecha.month]


class FuenteDatosExcel:
    # Esta clase solo sabe leer y preparar el Excel.

    def __init__(self, configuracion: ConfiguracionApp) -> None:
        # Guardamos la configuracion actual.
        self._configuracion = configuracion

    @property
    def configuracion(self) -> ConfiguracionApp:
        # Exponemos la configuracion actual en modo de solo lectura.
        return self._configuracion

    def cargar(self) -> pd.DataFrame:
        # Verificamos que el archivo configurado exista.
        if not self._configuracion.ruta_excel.exists():
            raise FileNotFoundError(
                f"No se encontro el archivo: {self._configuracion.ruta_excel}"
            )

        # Leemos la hoja indicada del Excel.
        marco_datos = pd.read_excel(
            self._configuracion.ruta_excel,
            sheet_name=self._configuracion.nombre_hoja,
            dtype=str,
        )

        # Normalizamos nombres de columnas para evitar errores por formato.
        marco_datos.columns = [
            str(columna).strip().upper() for columna in marco_datos.columns
        ]
        marco_datos["__INDICE_FUENTE__"] = range(1, len(marco_datos) + 1)

        # Validamos que exista la columna de serie.
        if self._configuracion.columna_serie not in marco_datos.columns:
            raise ValueError(
                f"No existe la columna '{self._configuracion.columna_serie}'"
            )

        # Validamos que exista la columna de RMA.
        if self._configuracion.columna_rma not in marco_datos.columns:
            raise ValueError(
                f"No existe la columna '{self._configuracion.columna_rma}'"
            )

        # Separamos celdas con varias series para que cada una pueda buscarse individualmente.
        marco_datos[self._configuracion.columna_serie] = (
            marco_datos[self._configuracion.columna_serie]
            .fillna("")
            .apply(extraer_series)
        )
        marco_datos = marco_datos.explode(
            self._configuracion.columna_serie,
            ignore_index=True,
        )
        marco_datos[self._configuracion.columna_serie] = (
            marco_datos[self._configuracion.columna_serie]
            .fillna("")
            .astype(str)
            .str.strip()
        )
        marco_datos = marco_datos[
            marco_datos[self._configuracion.columna_serie] != ""
        ].reset_index(drop=True)

        # Limpiamos la columna de RMA.
        marco_datos[self._configuracion.columna_rma] = (
            marco_datos[self._configuracion.columna_rma]
            .fillna("")
            .astype(str)
            .str.strip()
        )

        # Regresamos la tabla ya preparada.
        return marco_datos


class ServicioBusquedaRma:
    # Esta clase contiene la logica principal de busqueda.

    def __init__(
        self,
        configuracion: ConfiguracionApp,
        marco_datos: pd.DataFrame | None = None,
    ) -> None:
        # Guardamos la configuracion vigente.
        self._configuracion = configuracion
        # Guardamos la tabla cargada en memoria, si ya existe.
        self._marco_datos = marco_datos

    def actualizar_configuracion(self, configuracion: ConfiguracionApp) -> None:
        # Reemplazamos la configuracion anterior por la nueva.
        self._configuracion = configuracion

    def establecer_marco_datos(self, marco_datos: pd.DataFrame) -> None:
        # Guardamos los datos cargados para futuras consultas.
        self._marco_datos = marco_datos

    def tiene_datos(self) -> bool:
        # Indicamos si ya existe un DataFrame en memoria.
        return self._marco_datos is not None

    @property
    def marco_datos(self) -> pd.DataFrame | None:
        # Exponemos los datos en memoria para servicios de solo lectura.
        return self._marco_datos

    @property
    def configuracion(self) -> ConfiguracionApp:
        # Exponemos la configuracion vigente del servicio.
        return self._configuracion

    def buscar(self, serie: str) -> ResultadoBusqueda:
        # Si aun no hay datos cargados, lanzamos un error claro.
        if self._marco_datos is None:
            raise RuntimeError("Primero debes cargar un archivo de Excel.")

        # Normalizamos la serie para compararla correctamente.
        serie_limpia = normalizar_texto(serie)

        # Filtramos la tabla buscando coincidencias exactas.
        coincidencias = self._marco_datos[
            self._marco_datos[self._configuracion.columna_serie] == serie_limpia
        ]

        # Si no hay coincidencias, devolvemos estado SIN_RMA.
        if coincidencias.empty:
            return ResultadoBusqueda(estado=EstadoBusqueda.SIN_RMA, serie=serie_limpia)

        # Tomamos el valor de RMA de la primera coincidencia encontrada.
        fila_encontrada = coincidencias.iloc[0]
        rma = limpiar_valor_excel(fila_encontrada[self._configuracion.columna_rma])
        rma_age = self._obtener_valor_fila(fila_encontrada, "RMA AGE")
        condicion_rma = self._obtener_valor_fila(fila_encontrada, "CONDICION DEL RMA")

        # Normalizamos el valor del RMA para compararlo sin errores.
        rma_normalizado = normalizar_texto(rma)

        # Si el texto es SINCONTRATO, regresamos ese estado.
        if rma_normalizado == "SINCONTRATO":
            return ResultadoBusqueda(
                estado=EstadoBusqueda.SIN_CONTRATO,
                serie=serie_limpia,
                rma=rma,
                rma_age=rma_age,
                condicion_rma=condicion_rma,
            )

        # Si existe un valor y no es N/A, se considera un RMA valido.
        if rma and rma_normalizado != "N/A":
            series_relacionadas = self._obtener_series_por_rma(rma)
            return ResultadoBusqueda(
                estado=EstadoBusqueda.CON_RMA,
                serie=serie_limpia,
                rma=rma,
                rma_age=rma_age,
                condicion_rma=condicion_rma,
                series_relacionadas=tuple(series_relacionadas),
            )

        # En cualquier otro caso devolvemos ausencia de RMA utilizable.
        return ResultadoBusqueda(estado=EstadoBusqueda.SIN_RMA, serie=serie_limpia)

    def _obtener_valor_fila(self, fila, nombre_columna: str) -> str:
        # Buscamos columnas aunque vengan con acento, saltos o espacios distintos.
        objetivo = normalizar_nombre_columna(nombre_columna)
        for columna in fila.index:
            if normalizar_nombre_columna(columna) == objetivo:
                return limpiar_valor_excel(fila[columna])
        return ""

    def _obtener_series_por_rma(self, rma: str) -> list[str]:
        # Recuperamos todas las series que comparten el mismo RMA.
        if self._marco_datos is None:
            return []

        rma_limpio = str(rma).strip()
        if not rma_limpio:
            return []

        coincidencias = self._marco_datos[
            self._marco_datos[self._configuracion.columna_rma] == rma_limpio
        ]
        if coincidencias.empty:
            return []

        series = [
            str(valor).strip()
            for valor in coincidencias[self._configuracion.columna_serie].tolist()
            if str(valor).strip()
        ]

        # Conservamos el orden original sin repetir series.
        return list(dict.fromkeys(series))

    def marcar_rmas_completos_en_excel(self, rmas_completos: list[str]) -> Path:
        # Marcamos en verde las filas originales del Excel que pertenecen a RMAs completos.
        if self._marco_datos is None or self._marco_datos.empty:
            raise RuntimeError("Primero debes cargar un archivo Excel con datos validos.")

        rmas_limpios = {
            str(rma).strip()
            for rma in rmas_completos
            if str(rma).strip()
        }
        if not rmas_limpios:
            raise ValueError("No hay RMAs completos para marcar en el Excel.")

        ruta_excel = self._configuracion.ruta_excel
        if not ruta_excel.exists():
            raise FileNotFoundError(f"No se encontro el archivo: {ruta_excel}")

        filas_originales = self._obtener_filas_originales_por_rma(rmas_limpios)
        if not filas_originales:
            raise RuntimeError("No se encontraron filas originales para los RMAs completos.")

        libro = load_workbook(ruta_excel)
        if self._configuracion.nombre_hoja not in libro.sheetnames:
            raise ValueError(f"No existe la hoja '{self._configuracion.nombre_hoja}'.")

        hoja = libro[self._configuracion.nombre_hoja]
        relleno_completo = PatternFill(fill_type="solid", fgColor="C6EFCE")

        for fila_original in filas_originales:
            fila_excel = fila_original + 1
            for columna in range(1, hoja.max_column + 1):
                hoja.cell(row=fila_excel, column=columna).fill = relleno_completo

        libro.save(ruta_excel)
        return ruta_excel

    def _obtener_filas_originales_por_rma(self, rmas: set[str]) -> list[int]:
        if self._marco_datos is None or "__INDICE_FUENTE__" not in self._marco_datos.columns:
            return []

        coincidencias = self._marco_datos[
            self._marco_datos[self._configuracion.columna_rma].isin(rmas)
        ]
        filas: list[int] = []
        for valor in coincidencias["__INDICE_FUENTE__"].tolist():
            try:
                filas.append(int(valor))
            except (TypeError, ValueError):
                continue

        return sorted(set(filas))


class ServicioCargaExcel:
    # Esta clase coordina la carga del archivo y la sincronizacion de servicios.

    def __init__(
        self,
        fuente_datos: FuenteDatosExcel,
        servicio_busqueda: ServicioBusquedaRma,
    ) -> None:
        # Guardamos la fuente que sabe leer el Excel.
        self._fuente_datos = fuente_datos
        # Guardamos el servicio que realizara las consultas.
        self._servicio_busqueda = servicio_busqueda

    @property
    def configuracion_actual(self) -> ConfiguracionApp:
        # Regresamos la configuracion vigente de la fuente de datos.
        return self._fuente_datos.configuracion

    def cargar_desde_ruta(self, ruta_excel: Path) -> pd.DataFrame:
        # Creamos una nueva configuracion con la ruta actualizada.
        configuracion_actualizada = replace(
            self._fuente_datos.configuracion,
            ruta_excel=ruta_excel,
        )

        # Creamos una nueva fuente usando la nueva ruta.
        self._fuente_datos = FuenteDatosExcel(configuracion_actualizada)

        # Actualizamos la configuracion del servicio de busqueda.
        self._servicio_busqueda.actualizar_configuracion(configuracion_actualizada)

        # Cargamos el archivo desde la nueva ruta.
        marco_datos = self._fuente_datos.cargar()

        # Guardamos la tabla para que pueda consultarse.
        self._servicio_busqueda.establecer_marco_datos(marco_datos)

        # Regresamos la tabla cargada.
        return marco_datos

    def cargar_predeterminado(self) -> pd.DataFrame:
        # Cargamos el archivo definido en la configuracion inicial.
        marco_datos = self._fuente_datos.cargar()

        # Guardamos la tabla en el servicio de busqueda.
        self._servicio_busqueda.establecer_marco_datos(marco_datos)

        # Regresamos la tabla ya disponible para usar.
        return marco_datos


class GestorRmasSemanales:
    # Esta clase administra los RMAs semanales capturados manualmente.

    def __init__(self) -> None:
        # Lista interna donde se guardan los RMAs agregados durante la sesion.
        self._rmas_guardados: list[dict[str, str]] = []

    def guardar_desde_texto(self, texto: str) -> list[dict[str, str]]:
        # Separamos el texto por lineas y limpiamos espacios vacios.
        nuevas_entradas = [linea.strip() for linea in texto.splitlines() if linea.strip()]

        # Si no hubo entradas validas, regresamos lista vacia.
        if not nuevas_entradas:
            return []

        # Tomamos la fecha y hora actual una sola vez para las entradas de este guardado.
        fecha_actual = datetime.now()
        nombres_dias = {
            0: "Lunes",
            1: "Martes",
            2: "Miercoles",
            3: "Jueves",
            4: "Viernes",
            5: "Sabado",
            6: "Domingo",
        }
        dia_semana = nombres_dias[fecha_actual.weekday()]
        fecha_formateada = fecha_actual.strftime("%d/%m/%Y")
        hora_formateada = fecha_actual.strftime("%I:%M %p").lstrip("0")
        hora_formateada = (
            hora_formateada.replace("AM", "a. m.").replace("PM", "p. m.")
        )

        # Construimos el bloque de registros con su fecha asociada.
        registros_nuevos = [
            {
                "rma": entrada,
                "dia": dia_semana,
                "fecha": fecha_formateada,
                "hora": hora_formateada,
            }
            for entrada in nuevas_entradas
        ]

        # Agregamos las nuevas entradas a la coleccion acumulada.
        self._rmas_guardados.extend(registros_nuevos)

        # Regresamos lo que realmente se agrego para informar a la interfaz.
        return registros_nuevos

    def obtener_todos(self) -> list[dict[str, str]]:
        # Regresamos una copia para evitar modificaciones externas.
        return list(self._rmas_guardados)

    def tiene_rmas(self) -> bool:
        # Indicamos si ya hay al menos un RMA guardado.
        return len(self._rmas_guardados) > 0


class GeneradorVaciadoRma:
    # Construye el archivo de vaciado a partir del pending cargado y los RMA completos.

    COLUMNAS_VACIADO = [
        "MES",
        "AÑO",
        "MULTIVENDOR",
        "FECHA DE ENTREGA",
        "ACUSE",
        "NOMBRE DE QUIEN ENTREGA",
        "RMA",
        "TAREA",
        "PARTE INSTALADA",
        "SERIE INSTALADA",
        "PARTE RETORNADA",
        "SERIE RETORNADA",
        "DESCRIPCION DE PARTE",
        "ESTATUS",
    ]

    def __init__(
        self,
        configuracion: ConfiguracionApp,
        servicio_busqueda: ServicioBusquedaRma,
    ) -> None:
        self._configuracion = configuracion
        self._servicio_busqueda = servicio_busqueda

    def generar_desde_rmas_completos(
        self,
        rmas_completos: list[str],
        fecha_recoleccion: datetime | None = None,
        folio_cisco: str = "",
    ) -> Path:
        marco_datos = self._servicio_busqueda.marco_datos
        if marco_datos is None or marco_datos.empty:
            raise RuntimeError("Primero debes cargar un archivo Excel con datos validos.")

        if not rmas_completos:
            raise ValueError("No hay RMAs completos en la sesion para generar el vaciado.")

        filas_vaciado = self._construir_filas_vaciado(marco_datos, rmas_completos, fecha_recoleccion, folio_cisco)
        if not filas_vaciado:
            raise RuntimeError("No se pudo construir ninguna fila valida para el vaciado.")

        ruta_salida = self._construir_ruta_salida(folio_cisco)
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Pre alerta"

        self._poblar_hoja(hoja, filas_vaciado, folio_cisco)
        libro.save(ruta_salida)
        return ruta_salida

    def _construir_filas_vaciado(
        self,
        marco_datos: pd.DataFrame,
        rmas_completos: list[str],
        fecha_recoleccion: datetime | None = None,
        folio_cisco: str = "",
    ) -> list[dict[str, str]]:
        filas_vaciado: list[dict[str, str]] = []
        datos_rma = marco_datos[marco_datos[self._configuracion.columna_rma].isin(rmas_completos)]
        if datos_rma.empty:
            return []

        # Si el usuario proporcionó fecha manual, la usamos para todos los grupos
        if fecha_recoleccion is not None:
            # Usar la fecha seleccionada por el usuario para TODOS los RMAs
            fecha_entrega = fecha_recoleccion
        else:
            # Tomar fecha del SHIP DATE del primer grupo
            primera_fila = datos_rma.iloc[0]
            fecha_entrega = parsear_fecha_texto(primera_fila.get("SHIP DATE"))
            if fecha_entrega is None:
                fecha_entrega = datetime.now()

        grupos = datos_rma.groupby("__INDICE_FUENTE__", sort=False)
        for _indice, grupo in grupos:
            fila_base = grupo.iloc[0]

            # Si hay fecha manual, la usamos para todos los grupos
            # Si NO hay fecha manual, intentamos usar SHIP DATE por grupo
            if fecha_recoleccion is not None:
                mes_entrega = fecha_recoleccion
            else:
                mes_entrega = fecha_entrega

            series_retorno = [
                str(valor).strip().upper()
                for valor in grupo[self._configuracion.columna_serie].tolist()
                if str(valor).strip()
            ]
            series_instaladas = extraer_series(fila_base.get(self._configuracion.columna_numero, ""))
            while len(series_instaladas) < len(series_retorno):
                series_instaladas.append("")

            multivendor = str(fila_base.get("MV  Y/O RESPONSABLE", "TELEDINAMICA")).strip() or "TELEDINAMICA"
            # Construir acuse con formato "20 MAYO 2026" (sin folio, solo fecha)
            dia_formato = mes_entrega.strftime('%d')
            mes_formato = nombre_mes_espanol(mes_entrega)
            acuse = f"ACUSE CISCO {multivendor} {dia_formato} {mes_formato} {mes_entrega.strftime('%Y')}"
            parte = str(fila_base.get("PART NUMBER", "")).strip()
            rma = str(fila_base.get("RMA", "")).strip()
            tarea = str(fila_base.get("CUSTOMER REFERENCE #", "")).strip()
            recibe = self._configuracion.google_rma_recibe

            for indice, serie_vieja in enumerate(series_retorno):
                filas_vaciado.append(
                    {
                        "MES": nombre_mes_espanol(mes_entrega),
                        "AÑO": mes_entrega.strftime("%Y"),
                        "MULTIVENDOR": multivendor,
                        "FECHA DE ENTREGA": mes_entrega.strftime("%d/%m/%Y"),
                        "ACUSE": acuse,
                        "NOMBRE DE QUIEN ENTREGA": "ALFREDO CRUZ",
                        "RMA": rma,
                        "TAREA": tarea,
                        "PARTE INSTALADA": parte,
                        "SERIE INSTALADA": series_instaladas[indice],
                        "PARTE RETORNADA": parte,
                        "SERIE RETORNADA": serie_vieja,
                        "DESCRIPCION DE PARTE": "TELEFONO",
                        "ESTATUS": "DAÑADA",
                    }
                )

        return filas_vaciado

    def _construir_ruta_salida(self, folio_cisco: str = "") -> Path:
        carpeta_destino = self._servicio_busqueda.configuracion.ruta_excel.parent
        if folio_cisco:
            return carpeta_destino / f"ACUSE CISCO TELEDINAMICA {folio_cisco}.xlsx"
        else:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            return carpeta_destino / f"VACIADO_RMA_{timestamp}.xlsx"

    def _poblar_hoja(self, hoja, filas_vaciado: list[dict[str, str]], folio_cisco: str = "") -> None:
        relleno_encabezado = PatternFill(fill_type="solid", fgColor="0B2E59")
        fuente_encabezado = Font(color="FFFF00", bold=True)
        relleno_folio = PatternFill(fill_type="solid", fgColor="217346")  # Verde Microsoft
        fuente_folio = Font(color="FFFFFF", bold=True, size=16)
        borde = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        alineacion = Alignment(horizontal="center", vertical="center")

        for columna, titulo in enumerate(self.COLUMNAS_VACIADO, start=1):
            celda = hoja.cell(row=1, column=columna, value=titulo)
            celda.fill = relleno_encabezado
            celda.font = fuente_encabezado
            celda.border = borde
            celda.alignment = alineacion

        for fila_excel, fila in enumerate(filas_vaciado, start=2):
            for columna, titulo in enumerate(self.COLUMNAS_VACIADO, start=1):
                celda = hoja.cell(row=fila_excel, column=columna, value=fila[titulo])
                celda.border = borde
                celda.alignment = alineacion

        anchos = {
            1: 10,
            2: 8,
            3: 18,
            4: 16,
            5: 28,
            6: 24,
            7: 14,
            8: 20,
            9: 16,
            10: 18,
            11: 16,
            12: 18,
            13: 18,
            14: 14,
        }
        for indice, ancho in anchos.items():
            hoja.column_dimensions[get_column_letter(indice)].width = ancho

        fila_firmas = len(filas_vaciado) + 4
        hoja.merge_cells(start_row=fila_firmas, start_column=4, end_row=fila_firmas, end_column=6)
        hoja.merge_cells(start_row=fila_firmas + 1, start_column=4, end_row=fila_firmas + 1, end_column=6)
        hoja.merge_cells(start_row=fila_firmas, start_column=10, end_row=fila_firmas, end_column=12)
        hoja.merge_cells(start_row=fila_firmas + 1, start_column=10, end_row=fila_firmas + 1, end_column=12)

        hoja.cell(row=fila_firmas, column=4, value="PERSONAL QUE ENTREGA").alignment = alineacion
        hoja.cell(row=fila_firmas + 1, column=4, value="ALFREDO CRUZ").alignment = alineacion
        hoja.cell(row=fila_firmas, column=10, value="PERSONAL QUE RECIBE").alignment = alineacion
        hoja.cell(row=fila_firmas + 1, column=10, value="NOMBRE FECHA Y FIRMA").alignment = alineacion

        for fila in range(fila_firmas, fila_firmas + 2):
            for columna in (4, 5, 6, 10, 11, 12):
                hoja.cell(row=fila, column=columna).border = Border(bottom=Side(style="thin", color="000000"))

        # Fila del folio Cisco en verde DESPUES de las firmas
        if folio_cisco:
            hoja.merge_cells(start_row=fila_firmas + 4, start_column=1, end_row=fila_firmas + 4, end_column=14)
            celda_folio = hoja.cell(row=fila_firmas + 4, column=1, value=f"FOLIO: {folio_cisco}")
            celda_folio.fill = relleno_folio
            celda_folio.font = fuente_folio
            celda_folio.alignment = Alignment(horizontal="center", vertical="center")
            celda_folio.border = borde
            # Aumentar altura de la fila del folio
            hoja.row_dimensions[fila_firmas + 4].height = 35
